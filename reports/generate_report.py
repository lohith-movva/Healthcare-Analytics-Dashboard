"""
Automated Healthcare Analytics Report Generator
-------------------------------------------------
Reads from PostgreSQL (or CSV fallback), computes KPIs,
and generates a formatted Excel report — replacing manual refresh.

Author: Lohith Movva
"""

import pandas as pd
import numpy as np
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime
import os
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

DB_URL = os.getenv("DATABASE_URL", "postgresql://postgres:password@localhost:5432/healthcare_analytics")
OUTPUT_DIR = "reports"
os.makedirs(OUTPUT_DIR, exist_ok=True)

NAVY  = "1A2E4A"
WHITE = "FFFFFF"
LIGHT = "EEF2F8"
GREEN = "1A6B38"
RED   = "C0392B"

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_data():
    try:
        engine = create_engine(DB_URL)
        logger.info("Connected to PostgreSQL.")
        summary   = pd.read_sql("SELECT * FROM kpi_monthly_summary ORDER BY month", engine)
        encounters = pd.read_sql("SELECT * FROM fact_encounters", engine)
        claims     = pd.read_sql("SELECT * FROM fact_claims", engine)
        patients   = pd.read_sql("SELECT * FROM dim_patients", engine)
    except Exception as e:
        logger.warning(f"DB unavailable ({e}), loading from CSV...")
        summary    = pd.read_csv("data/kpi_monthly_summary.csv")
        encounters = pd.read_csv("data/fact_encounters.csv")
        claims     = pd.read_csv("data/fact_claims.csv")
        patients   = pd.read_csv("data/dim_patients.csv")
    return summary, encounters, claims, patients


# ---------------------------------------------------------------------------
# KPI calculations
# ---------------------------------------------------------------------------

def compute_kpis(summary, encounters, claims, patients):
    return {
        "total_patients":       len(patients),
        "total_encounters":     len(encounters),
        "total_claims":         len(claims),
        "avg_los":              round(encounters["length_of_stay"].mean(), 2),
        "readmission_rate":     round(encounters["readmitted_30d"].mean() * 100, 2),
        "avg_cost_encounter":   round(encounters["cost_per_encounter"].mean(), 2),
        "total_billed":         round(claims["claim_amount"].sum(), 2),
        "total_approved":       round(claims["approved_amount"].sum(), 2),
        "denial_rate":          round((~claims["approved"]).mean() * 100, 2),
        "chronic_patient_pct":  round(patients["is_chronic"].mean() * 100, 2) if "is_chronic" in patients else 0,
    }


# ---------------------------------------------------------------------------
# Excel report builder
# ---------------------------------------------------------------------------

def header_style(ws, row, col, value, width=18):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def data_cell(ws, row, col, value, number_format=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER
    if number_format:
        cell.number_format = number_format
    if row % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=LIGHT)
    return cell


def build_kpi_sheet(wb, kpis):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "Healthcare Analytics — Executive Summary"
    title.font = Font(bold=True, size=14, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"Report generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
    sub.font = Font(italic=True, size=10, color="555555")
    sub.alignment = Alignment(horizontal="center")

    headers = ["KPI", "Value", "Benchmark", "Status", "Notes"]
    for i, h in enumerate(headers, 1):
        header_style(ws, 4, i, h, width=22)

    rows = [
        ("Total Patients",              f"{kpis['total_patients']:,}",        "—",        "—",        "Active patient population"),
        ("Total Encounters",            f"{kpis['total_encounters']:,}",       "—",        "—",        "All inpatient & outpatient visits"),
        ("Total Claims",                f"{kpis['total_claims']:,}",          "—",        "—",        "Submitted to payers"),
        ("Avg Length of Stay (days)",   f"{kpis['avg_los']}",                 "4.5",      "✓ Good" if kpis["avg_los"] <= 4.5 else "⚠ High", "National avg ~4.5 days"),
        ("30-Day Readmission Rate",     f"{kpis['readmission_rate']}%",       "15%",      "✓ Good" if kpis["readmission_rate"] <= 15 else "⚠ High", "CMS benchmark ≤15%"),
        ("Avg Cost per Encounter",      f"${kpis['avg_cost_encounter']:,.0f}","$4,000",   "—",        "Blended avg across service lines"),
        ("Total Billed",                f"${kpis['total_billed']:,.0f}",      "—",        "—",        "Gross charges"),
        ("Total Approved",              f"${kpis['total_approved']:,.0f}",    "—",        "—",        "Net collections"),
        ("Claims Denial Rate",          f"{kpis['denial_rate']}%",            "5%",       "✓ Good" if kpis["denial_rate"] <= 5 else "⚠ High", "Target <5%"),
        ("Chronic Patient %",           f"{kpis['chronic_patient_pct']}%",    "—",        "—",        "Diabetes/HTN/CHF/COPD"),
    ]

    for r_idx, row_data in enumerate(rows, 5):
        for c_idx, val in enumerate(row_data, 1):
            cell = data_cell(ws, r_idx, c_idx, val, bold=(c_idx == 1))
            if c_idx == 4:
                if "✓" in str(val):
                    cell.font = Font(color=GREEN, bold=True, size=10)
                elif "⚠" in str(val):
                    cell.font = Font(color=RED, bold=True, size=10)

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 26

    ws.row_dimensions[4].height = 22
    return ws


def build_monthly_sheet(wb, summary):
    ws = wb.create_sheet("Monthly Trends")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = "Monthly KPI Trends"
    title.font = Font(bold=True, size=13, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = [
        ("Month", "month"),
        ("Encounters", "total_encounters"),
        ("Claims", "total_claims"),
        ("Avg LOS", "avg_los"),
        ("Readmissions", "readmission_count"),
        ("Readmit Rate %", "readmission_rate"),
        ("Total Billed", "total_billed"),
        ("Total Approved", "total_approved"),
        ("Denial Rate %", "denial_rate_pct"),
    ]

    for c_idx, (label, _) in enumerate(cols, 1):
        header_style(ws, 3, c_idx, label, width=18)

    for r_idx, (_, row) in enumerate(summary.iterrows(), 4):
        for c_idx, (_, field) in enumerate(cols, 1):
            val = row.get(field, "")
            if field in ("total_billed", "total_approved"):
                data_cell(ws, r_idx, c_idx, val, number_format='"$"#,##0.00')
            elif field in ("readmission_rate", "denial_rate_pct"):
                data_cell(ws, r_idx, c_idx, round(float(val) * 100, 2) if val else 0, number_format='0.00"%"')
            else:
                data_cell(ws, r_idx, c_idx, val)

    # Line chart — readmission rate
    if len(summary) > 1:
        chart = LineChart()
        chart.title = "Monthly Readmission Rate"
        chart.style = 10
        chart.y_axis.title = "Rate"
        chart.x_axis.title = "Month"
        data_ref = Reference(ws, min_col=6, min_row=3, max_row=3 + len(summary))
        chart.add_data(data_ref, titles_from_data=True)
        chart.width = 20
        chart.height = 12
        ws.add_chart(chart, "A" + str(len(summary) + 6))

    return ws


def build_dept_sheet(wb, encounters):
    ws = wb.create_sheet("Department Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Encounter Metrics by Department"
    t.font = Font(bold=True, size=13, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    dept_summary = (
        encounters.groupby("department")
        .agg(
            total_encounters=("encounter_id", "count"),
            avg_los=("length_of_stay", "mean"),
            readmission_rate=("readmitted_30d", "mean"),
            avg_cost=("cost_per_encounter", "mean"),
            total_cost=("cost_per_encounter", "sum"),
        )
        .round(2)
        .reset_index()
        .sort_values("total_encounters", ascending=False)
    )

    headers = ["Department", "Encounters", "Avg LOS", "Readmit Rate %", "Avg Cost", "Total Cost"]
    for c, h in enumerate(headers, 1):
        header_style(ws, 3, c, h, width=20)

    for r_idx, (_, row) in enumerate(dept_summary.iterrows(), 4):
        data_cell(ws, r_idx, 1, row["department"], bold=True)
        data_cell(ws, r_idx, 2, int(row["total_encounters"]))
        data_cell(ws, r_idx, 3, row["avg_los"])
        data_cell(ws, r_idx, 4, round(row["readmission_rate"] * 100, 2))
        data_cell(ws, r_idx, 5, row["avg_cost"], number_format='"$"#,##0.00')
        data_cell(ws, r_idx, 6, row["total_cost"], number_format='"$"#,##0.00')

    # Bar chart — encounters by dept
    chart = BarChart()
    chart.type = "col"
    chart.title = "Encounters by Department"
    chart.y_axis.title = "Count"
    chart.style = 10
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + len(dept_summary))
    cats   = Reference(ws, min_col=1, min_row=4, max_row=3 + len(dept_summary))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 12
    ws.add_chart(chart, "A" + str(len(dept_summary) + 6))

    return ws


def generate_report():
    logger.info("Loading data...")
    summary, encounters, claims, patients = load_data()
    kpis = compute_kpis(summary, encounters, claims, patients)

    logger.info("Building Excel report...")
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_kpi_sheet(wb, kpis)
    build_monthly_sheet(wb, summary)
    build_dept_sheet(wb, encounters)

    filename = f"Healthcare_Analytics_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    logger.info(f"Report saved → {filepath}")
    return filepath


if __name__ == "__main__":
    generate_report()
